# Copyright (c) 2026, DAS and contributors
# For license information, please see license.txt
import os

import pandas as pd
from openpyxl import load_workbook

import frappe
from frappe.model.document import Document

from sth.utils import generate_duplicate_key

class PaymenttoExcel(Document):

	def before_submit(self):
		for d in self.items:
			generate_duplicate_key(d, "duplicate_key", [self.voucher_type, self.voucher_no])

	def on_submit(self):
		self.update_payment_entry_status()
		self.create_excel_file()

	def before_cancel(self):
		for d in self.items:
			generate_duplicate_key(d, "duplicate_key", cancel=1)

	def on_cancel(self):
		self.remove_excel_file()
		self.update_payment_entry_status()

	def remove_excel_file(self):
		# Hapus file jika ada
		path = f'{frappe.get_site_path()}{self.exported_excel}'
		if os.path.exists(path):
			os.remove(path)

	def update_payment_entry_status(self):
		pass
	
	def create_excel_file(self):
		self.exported_excel = f'/private/mandiri/{self.name}.xlsx'

		data = self.set_data_payment()
		self.process_data_to_excel(data)

	def set_data_payment(self):
		return {'ID': {0: 23, 1: 43, 2: 12,
						3: 13, 4: 67, 5: 89,
						6: 90, 7: 56, 8: 34},
				'Name': {0: 'Ram', 1: 'Deep',
						2: 'Yash', 3: 'Aman',
						4: 'Arjun', 5: 'Aditya',
						6: 'Divya', 7: 'Chalsea',
						8: 'Akash'},
				'Marks': {0: 89, 1: 97, 2: 45, 3: 78,
							4: 56, 5: 76, 6: 100, 7: 87,
							8: 81},
				'Grade': {0: 'B', 1: 'A', 2: 'F', 3: 'C',
							4: 'E', 5: 'C', 6: 'A', 7: 'B',
							8: 'B'}}
	
	def process_data_to_excel(self, data):
		df = pd.DataFrame(data)

		writer = pd.ExcelWriter(f'{frappe.get_site_path()}{self.exported_excel}', engine='openpyxl')

		df.to_excel(writer, index=False)  # Mulai dari baris ke-3 untuk data

		writer.close()
	# 	worksheet = writer.sheets['Data']

	# 	# Tambahkan dua baris di atas data
	# 	worksheet.insert_rows(0, amount=1)

	# 	# Isi baris pertama dengan judul "Laporan Harian Stock - Sales"
	# 	worksheet.cell(row=1, column=1, value=f"Laporan Harian Stock - Sales (Atas {'Cabang '+str(list(data_cabang.keys())[-1]) if len(data_cabang) == 2 else 'Semua Cabang'})")
	# 	worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)

	# 	# Isi baris kedua dengan informasi periode
	# 	worksheet.cell(row=2, column=1, value=f"Periode Bulan : {period_list[0]['label']} s/d {period_list[-1]['label']}")
	# 	worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

	# 	# Isi header untuk data
	# 	worksheet.cell(row=4, column=1, value="Kode")
	# 	worksheet.merge_cells(start_row=4, start_column=1, end_row=5, end_column=1)

	# 	worksheet.cell(row=4, column=2, value="Nama Item")
	# 	worksheet.merge_cells(start_row=4, start_column=2, end_row=5, end_column=2)

	# 	idx_start = 3
	# 	idx_end = 3 + (11 + (len(period_list) * 5))
	# 	for key, valuenya in data_cabang.items():
	# 		worksheet.cell(row=4, column=idx_start, value=valuenya)
	# 		worksheet.merge_cells(start_row=4, start_column=idx_start, end_row=4, end_column=idx_end)

	# 		idx_start = idx_end + 1
	# 		idx_end = idx_start + (10 + (len(period_list) * 5)) + 1
		
	# 	writer.save()


	# 	# Load workbook
	# 	workbook = load_workbook(f'/home/frappe/frappe-bench/sites/{frappe.local.site}/private/files/Laporan Harian Stock.xlsx')

	# 	# Akses lembar kerja 'Data'
	# 	worksheet = workbook['Data']

	# 	# Atur lebar kolom 1 dan 2
	# 	worksheet.column_dimensions['A'].width = 30  # Lebar kolom A
	# 	worksheet.column_dimensions['B'].width = 50  # Lebar kolom B

	# 	# Simpan perubahan
	# 	workbook.save(f'/home/frappe/frappe-bench/sites/{frappe.local.site}/private/files/Laporan Harian Stock.xlsx')